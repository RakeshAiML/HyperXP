from io import BytesIO
from openpyxl import load_workbook
from excel_gen import generate_workbook

_SHEETS = [
    {
        "name": "Raw Materials",
        "columns": ["S.No", "Material Name", "UOM"],
        "rows": [
            {
                "S.No":          {"value": "1",     "confidence": "high"},
                "Material Name": {"value": "ETC-3", "confidence": "high"},
                "UOM":           {"value": "Kg",    "confidence": "low"},
            },
            {
                "S.No":          {"value": "2",    "confidence": "high"},
                "Material Name": {"value": None,   "confidence": "low"},
                "UOM":           {"value": "L",    "confidence": "high"},
            },
        ],
    },
    {
        "name": "Shift Log",
        "columns": ["Date", "Shift"],
        "rows": [
            {
                "Date":  {"value": "06/10/2025", "confidence": "high"},
                "Shift": {"value": "A",          "confidence": "high"},
            }
        ],
    },
]


def test_returns_bytes():
    result = generate_workbook(_SHEETS)
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_sheet_count_matches_input():
    wb = load_workbook(BytesIO(generate_workbook(_SHEETS)))
    assert len(wb.sheetnames) == 2


def test_sheet_names_match():
    wb = load_workbook(BytesIO(generate_workbook(_SHEETS)))
    assert "Raw Materials" in wb.sheetnames
    assert "Shift Log" in wb.sheetnames


def test_header_row_is_bold():
    wb = load_workbook(BytesIO(generate_workbook(_SHEETS)))
    ws = wb["Raw Materials"]
    assert ws.cell(row=1, column=1).font.bold is True


def test_header_row_frozen():
    wb = load_workbook(BytesIO(generate_workbook(_SHEETS)))
    ws = wb["Raw Materials"]
    assert ws.freeze_panes == "A2"


def test_column_headers_written():
    wb = load_workbook(BytesIO(generate_workbook(_SHEETS)))
    ws = wb["Raw Materials"]
    assert ws.cell(row=1, column=1).value == "S.No"
    assert ws.cell(row=1, column=2).value == "Material Name"
    assert ws.cell(row=1, column=3).value == "UOM"


def test_data_rows_written():
    wb = load_workbook(BytesIO(generate_workbook(_SHEETS)))
    ws = wb["Raw Materials"]
    assert ws.max_row == 3  # 1 header + 2 data rows


def test_low_confidence_cell_is_yellow():
    wb = load_workbook(BytesIO(generate_workbook(_SHEETS)))
    ws = wb["Raw Materials"]
    # Row 2, Col 3 = UOM "Kg" with confidence "low"
    fill = ws.cell(row=2, column=3).fill
    assert fill.fgColor.rgb == "00FFFF00"


def test_null_value_cell_is_red():
    wb = load_workbook(BytesIO(generate_workbook(_SHEETS)))
    ws = wb["Raw Materials"]
    # Row 3, Col 2 = Material Name null
    fill = ws.cell(row=3, column=2).fill
    assert fill.fgColor.rgb == "00FF6B6B"


def test_high_confidence_non_null_cell_not_highlighted():
    wb = load_workbook(BytesIO(generate_workbook(_SHEETS)))
    ws = wb["Raw Materials"]
    # Row 2, Col 1 = S.No "1" high confidence
    fill = ws.cell(row=2, column=1).fill
    assert fill.fgColor.rgb not in ("00FFFF00", "00FF6B6B")


def test_null_value_written_as_empty_string():
    wb = load_workbook(BytesIO(generate_workbook(_SHEETS)))
    ws = wb["Raw Materials"]
    # Row 3, Col 2 = null → empty cell
    assert ws.cell(row=3, column=2).value is None


def test_second_sheet_data_correct():
    wb = load_workbook(BytesIO(generate_workbook(_SHEETS)))
    ws = wb["Shift Log"]
    assert ws.cell(row=2, column=1).value == "06/10/2025"
    assert ws.cell(row=2, column=2).value == "A"
