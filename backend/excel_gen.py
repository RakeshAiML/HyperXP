from io import BytesIO
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

_YELLOW = PatternFill("solid", fgColor="FFFF00")
_RED    = PatternFill("solid", fgColor="FF6B6B")
_BOLD   = Font(bold=True)


def _auto_fit_columns(ws) -> None:
    for col in ws.columns:
        if not col:
            continue
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


def generate_workbook(sheets: List[dict]) -> bytes:
    """Create an Excel workbook with one sheet per entry in sheets[].

    Each sheet entry: { name, columns: [...], rows: [{col: {value, confidence}}] }
    Yellow fill = confidence "low" with non-null value.
    Red fill    = value is null.
    """
    if not sheets:
        raise ValueError("sheets must contain at least one entry")
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    for sheet_def in sheets:
        ws = wb.create_sheet(title=sheet_def["name"][:31])  # Excel tab name limit
        columns = sheet_def.get("columns", [])

        # Header row
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = _BOLD
        ws.freeze_panes = "A2"

        # Data rows
        for row_idx, row_data in enumerate(sheet_def.get("rows", []), 2):
            for col_idx, col_name in enumerate(columns, 1):
                cell_obj = row_data.get(col_name, {})
                value      = cell_obj.get("value")
                confidence = cell_obj.get("confidence", "high")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if value is None:
                    cell.fill = _RED
                elif confidence == "low":
                    cell.fill = _YELLOW

        _auto_fit_columns(ws)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
